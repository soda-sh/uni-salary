# resources to understand what's going on here:
    # https://youtu.be/7YS6YDQKFh0?si=N-D3giaDBtqcaV7y
    # https://openpyxl.readthedocs.io/en/stable/


import openpyxl as excel
import openpyxl.utils as utils
from output_dialog import Ui_Dialog_Output as q


# this class works like an API, contains methods to work with excel files
class Excel:
    # searches for a value inside a list formatted excel file, returns a row
    def listSearch(self, workbook, headerName, value):
        wb = excel.load_workbook(workbook)
        ws = wb.active

        # finding the column to search inside for a match
        for cell in ws[1]:
            if cell.value == headerName:
                colLetter = utils.get_column_letter(cell.column)

        # finding the matching row
        for cell in ws[colLetter]:
            if cell.value == value:
                rowNumber = cell.row

        # if there's a match, it returns the row, else it returns false
        try:
            returnList = ws[rowNumber]
        except UnboundLocalError:
            q.tprint("استاد در لیست وجود ندارد")
            returnList = False
        return returnList

    # searches for a value inside a table formatted excel file, returns the matching cell
    def tableSearch(self, workbook, worksheet, colHeader, rowHeader):
        wb = excel.load_workbook(workbook)

        # setting the active worksheet
        # to use the default worksheet, call the method with empty string '' as worksheet argument
        if worksheet == '':
            ws = wb.active
        else:
            ws = wb[worksheet]

        # finding the coordinates of the cell
        for cell in ws[1]:
            if cell.value == colHeader:
                colLetter = utils.get_column_letter(cell.column)

        for cell in ws[utils.get_column_letter(ws.max_column)]:
            if cell.value == rowHeader:
                rowNumber = cell.row

        # if the cell exists, it returns the cell, else it returns false
        # if you're getting UnboundLocalError, it means the search was unsuccessful
        # it can be caused by: wrong header names, non-existent value, etc.
        # you can change this to a try except block, like in listSearch()
        if ws[f'{colLetter}{rowNumber}']:
            return ws[f'{colLetter}{rowNumber}']
        else:
            return False

    # appends a list of values to a workbook (excel file)
    def append(self, workbook, values):
        wb = excel.load_workbook(workbook)
        ws = wb.active

        ws.append(values)
        wb.save(workbook)

    # creates an excel file, takes file name as argument (without extension: .xlsx)
    def createWorkbook(self, name):
        wb = excel.Workbook()
        wb.save(f"{name}.xlsx")
        # returns file path (with extension) to work with
        return f"{name}.xlsx"
